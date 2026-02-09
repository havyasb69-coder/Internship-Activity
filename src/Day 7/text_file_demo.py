import os
print(os.getcwd())

file = open("sample.txt", "w")
file.write("Hello,This is a sample text file.\n")
file.close()
file = open("sample.txt", "r")
content = file.read()   
print(content)
file.close()    

#using with statement
with open("sample.txt", "r") as file:
    content = file.read()
    print(content)

#Error handling with try-except
try:
    with open("missing.txt", "r") as file:
        content = file.read()
        print(content)
except FileNotFoundError:
    print("The file 'missing.txt' does not exist.") 

#CSV  parsing Basic
import csv

with open("data.csv", "r") as file:
    reader = csv.reader(file)

    for row in reader:
        print(row)
        print(row[0], row[1], row[2])

#Exel parsing Basic
import openpyxl 
from openpyxl import load_workbook

wb = load_workbook("data.xlsx")
sheet = wb.active

for row in sheet.iter_rows(values_only=True):
    print(row)
